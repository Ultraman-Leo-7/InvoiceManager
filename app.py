# -*- coding: utf-8 -*-
"""InvoiceManager v5.2.10 GUI."""

from __future__ import annotations

import base64
import ctypes
import json
import os
import sqlite3
import subprocess
import sys
import threading
import traceback
from ctypes import wintypes
from datetime import datetime
from pathlib import Path
import tkinter as tk
from tkinter import messagebox, ttk

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from data_backup import (
    LATEST_NAME,
    NutstoreWebDAV,
    create_local_safety_backup,
    create_portable_snapshot,
    restore_snapshot,
    temporary_snapshot_path,
)
from invoice_extract import (
    AVAILABLE_FIELDS,
    DEFAULT_FIELDS,
    PARSER_VERSION,
    parse_invoice,
    sha256_file,
)
from jd_qq import fetch_jd_invoices
from purchase_tracker import (
    add_purchase,
    clear_manual_match,
    clear_purchases,
    delete_purchase,
    get_purchase,
    init_purchase_table,
    match_purchases,
    required_components,
    set_manual_match,
    update_purchase,
)
from self_update import download_verified_update, latest_release, schedule_windows_replacement

APP_TITLE = "InvoiceManager"
APP_VERSION = "5.2.10"
DB_FILENAME = ".invoice_manager.db"
EXPORT_FILENAME = "发票汇总.xlsx"
FOLDER_POLL_MS = 2000
LOCAL_BACKUP_DIRNAME = ".invoice_manager_backups"


def app_dir() -> Path:
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent


BASE_DIR = app_dir()
DB_PATH = BASE_DIR / DB_FILENAME
EXPORT_PATH = BASE_DIR / EXPORT_FILENAME
LOCAL_BACKUP_DIR = BASE_DIR / LOCAL_BACKUP_DIRNAME


class DATA_BLOB(ctypes.Structure):
    _fields_ = [("cbData", wintypes.DWORD), ("pbData", ctypes.POINTER(ctypes.c_byte))]


def _blob_from_bytes(data: bytes):
    if not data:
        data = b"\x00"
    buf = ctypes.create_string_buffer(data, len(data))
    blob = DATA_BLOB(len(data), ctypes.cast(buf, ctypes.POINTER(ctypes.c_byte)))
    return blob, buf


def protect_secret(text: str) -> str:
    if not text:
        return ""
    raw = text.encode("utf-8")
    if os.name != "nt":
        return "plain:" + base64.b64encode(raw).decode("ascii")
    in_blob, in_buf = _blob_from_bytes(raw)
    out_blob = DATA_BLOB()
    crypt32 = ctypes.windll.crypt32
    kernel32 = ctypes.windll.kernel32
    kernel32.LocalFree.argtypes = [ctypes.c_void_p]
    kernel32.LocalFree.restype = ctypes.c_void_p
    ok = crypt32.CryptProtectData(ctypes.byref(in_blob), None, None, None, None, 0, ctypes.byref(out_blob))
    if not ok:
        raise ctypes.WinError()
    try:
        encrypted = ctypes.string_at(out_blob.pbData, out_blob.cbData)
        return "dpapi:" + base64.b64encode(encrypted).decode("ascii")
    finally:
        kernel32.LocalFree(out_blob.pbData)
        _ = in_buf


def unprotect_secret(value: str) -> str:
    if not value:
        return ""
    if value.startswith("plain:"):
        return base64.b64decode(value[6:]).decode("utf-8")
    if not value.startswith("dpapi:") or os.name != "nt":
        return ""
    encrypted = base64.b64decode(value[6:])
    in_blob, in_buf = _blob_from_bytes(encrypted)
    out_blob = DATA_BLOB()
    crypt32 = ctypes.windll.crypt32
    kernel32 = ctypes.windll.kernel32
    kernel32.LocalFree.argtypes = [ctypes.c_void_p]
    kernel32.LocalFree.restype = ctypes.c_void_p
    ok = crypt32.CryptUnprotectData(ctypes.byref(in_blob), None, None, None, None, 0, ctypes.byref(out_blob))
    if not ok:
        raise ctypes.WinError()
    try:
        return ctypes.string_at(out_blob.pbData, out_blob.cbData).decode("utf-8")
    finally:
        kernel32.LocalFree(out_blob.pbData)
        _ = in_buf


def hide_windows_file(path: Path) -> None:
    if os.name != "nt" or not path.exists():
        return
    attrs = ctypes.windll.kernel32.GetFileAttributesW(str(path))
    if attrs != 0xFFFFFFFF:
        ctypes.windll.kernel32.SetFileAttributesW(str(path), attrs | 0x2)


def connect_db() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA busy_timeout=5000")
    return conn


def get_setting_conn(conn: sqlite3.Connection, key: str):
    row = conn.execute("SELECT value FROM settings WHERE key=?", (key,)).fetchone()
    return row["value"] if row else None


def set_setting_conn(conn: sqlite3.Connection, key: str, value: str) -> None:
    conn.execute(
        "INSERT INTO settings(key, value) VALUES(?, ?) ON CONFLICT(key) DO UPDATE SET value=excluded.value",
        (key, value),
    )


def init_db() -> None:
    with connect_db() as conn:
        conn.executescript("""
            CREATE TABLE IF NOT EXISTS invoices (
                digest TEXT PRIMARY KEY,
                filename TEXT NOT NULL,
                invoice_no TEXT,
                issue_date TEXT,
                buyer TEXT,
                seller TEXT,
                project TEXT,
                total REAL,
                amount_wo_tax REAL,
                tax REAL,
                issuer TEXT,
                parse_status TEXT,
                parser_version INTEGER NOT NULL DEFAULT 0,
                confirmed INTEGER NOT NULL DEFAULT 0,
                note TEXT NOT NULL DEFAULT '',
                active INTEGER NOT NULL DEFAULT 1,
                last_seen TEXT NOT NULL
            );
            CREATE INDEX IF NOT EXISTS idx_invoices_filename ON invoices(filename);
            CREATE INDEX IF NOT EXISTS idx_invoices_invoice_no ON invoices(invoice_no);
            CREATE INDEX IF NOT EXISTS idx_invoices_active ON invoices(active);
            CREATE TABLE IF NOT EXISTS settings (key TEXT PRIMARY KEY, value TEXT NOT NULL);
        """)
        init_purchase_table(conn)
        if get_setting_conn(conn, "selected_fields") is None:
            fields = DEFAULT_FIELDS.copy()
            legacy = BASE_DIR / "settings.json"
            if legacy.exists():
                try:
                    data = json.loads(legacy.read_text(encoding="utf-8-sig"))
                    raw = data.get("fields")
                    if isinstance(raw, list):
                        fields = [x for x in raw if x in AVAILABLE_FIELDS] or fields
                except Exception:
                    pass
            set_setting_conn(conn, "selected_fields", json.dumps(fields, ensure_ascii=False))
    hide_windows_file(DB_PATH)


def get_setting(key: str, default: str = "") -> str:
    with connect_db() as conn:
        value = get_setting_conn(conn, key)
        return default if value is None else value


def set_setting(key: str, value: str) -> None:
    with connect_db() as conn:
        set_setting_conn(conn, key, value)


def selected_fields() -> list[str]:
    try:
        raw = json.loads(get_setting("selected_fields", "[]"))
    except Exception:
        raw = []
    result = [x for x in raw if x in AVAILABLE_FIELDS]
    return result or DEFAULT_FIELDS.copy()


FIELD_DB_MAP = {
    "文件名": "filename",
    "发票号码": "invoice_no",
    "开票日期": "issue_date",
    "购买方名称": "buyer",
    "销售方名称": "seller",
    "项目名称": "project",
    "价税合计": "total",
    "金额（不含税）": "amount_wo_tax",
    "税额": "tax",
    "开票人": "issuer",
    "提取状态": "parse_status",
}

SEARCH_SCOPES = [
    "全部字段", "文件名", "项目名称", "价税合计", "发票号码", "开票日期",
    "购买方名称", "销售方名称", "金额（不含税）", "税额", "开票人", "提取状态",
    "购买匹配", "人工确认", "备注",
]


def invoice_record_to_db(record: dict) -> dict:
    return {
        "filename": record.get("文件名", ""),
        "invoice_no": record.get("发票号码", ""),
        "issue_date": record.get("开票日期", ""),
        "buyer": record.get("购买方名称", ""),
        "seller": record.get("销售方名称", ""),
        "project": record.get("项目名称", ""),
        "total": record.get("价税合计"),
        "amount_wo_tax": record.get("金额（不含税）"),
        "tax": record.get("税额"),
        "issuer": record.get("开票人", ""),
        "parse_status": record.get("提取状态", ""),
    }


def sync_folder(progress=None) -> dict:
    pdfs = sorted(BASE_DIR.glob("*.pdf"), key=lambda p: p.name.lower())
    now = datetime.now().isoformat(timespec="seconds")
    stats = {"total": len(pdfs), "reused": 0, "parsed": 0, "failed": 0, "removed": 0}
    with connect_db() as conn:
        old_active = {row["digest"] for row in conn.execute("SELECT digest FROM invoices WHERE active=1")}
        conn.execute("UPDATE invoices SET active=0")
        for i, pdf in enumerate(pdfs, 1):
            if progress:
                progress(f"扫描 {i}/{len(pdfs)}：{pdf.name}")
            try:
                digest = sha256_file(pdf)
                row = conn.execute("SELECT * FROM invoices WHERE digest=?", (digest,)).fetchone()
                if row and int(row["parser_version"] or 0) == PARSER_VERSION:
                    conn.execute(
                        "UPDATE invoices SET filename=?, active=1, last_seen=? WHERE digest=?",
                        (pdf.name, now, digest),
                    )
                    stats["reused"] += 1
                    continue
                record = parse_invoice(pdf)
                data = invoice_record_to_db(record)
                manual = conn.execute(
                    "SELECT confirmed, note FROM invoices WHERE filename=? ORDER BY active DESC, last_seen DESC LIMIT 1",
                    (pdf.name,),
                ).fetchone()
                if not manual and data["invoice_no"]:
                    manual = conn.execute(
                        "SELECT confirmed, note FROM invoices WHERE invoice_no=? ORDER BY active DESC, last_seen DESC LIMIT 1",
                        (data["invoice_no"],),
                    ).fetchone()
                confirmed = int(manual["confirmed"]) if manual else 0
                note = str(manual["note"] or "") if manual else ""
                conn.execute("""
                    INSERT INTO invoices(
                        digest, filename, invoice_no, issue_date, buyer, seller, project, total,
                        amount_wo_tax, tax, issuer, parse_status, parser_version, confirmed, note,
                        active, last_seen
                    )
                    VALUES(
                        :digest, :filename, :invoice_no, :issue_date, :buyer, :seller, :project, :total,
                        :amount_wo_tax, :tax, :issuer, :parse_status, :parser_version, :confirmed, :note,
                        1, :last_seen
                    )
                    ON CONFLICT(digest) DO UPDATE SET
                        filename=excluded.filename,
                        invoice_no=excluded.invoice_no,
                        issue_date=excluded.issue_date,
                        buyer=excluded.buyer,
                        seller=excluded.seller,
                        project=excluded.project,
                        total=excluded.total,
                        amount_wo_tax=excluded.amount_wo_tax,
                        tax=excluded.tax,
                        issuer=excluded.issuer,
                        parse_status=excluded.parse_status,
                        parser_version=excluded.parser_version,
                        active=1,
                        last_seen=excluded.last_seen
                """, {
                    "digest": digest,
                    "parser_version": PARSER_VERSION,
                    "confirmed": confirmed,
                    "note": note,
                    "last_seen": now,
                    **data,
                })
                stats["parsed"] += 1
            except Exception as e:
                stats["failed"] += 1
                if progress:
                    progress(f"解析失败：{pdf.name} | {type(e).__name__}: {e}")
        current_active = {row["digest"] for row in conn.execute("SELECT digest FROM invoices WHERE active=1")}
        stats["removed"] = len(old_active - current_active)
    return stats


def all_active_rows():
    with connect_db() as conn:
        return conn.execute("SELECT * FROM invoices WHERE active=1 ORDER BY filename COLLATE NOCASE").fetchall()


def set_confirmed_many(digests: list[str], value: bool) -> None:
    if not digests:
        return
    with connect_db() as conn:
        conn.executemany(
            "UPDATE invoices SET confirmed=? WHERE digest=?",
            [(1 if value else 0, d) for d in digests],
        )


def set_note(digest: str, note: str) -> None:
    with connect_db() as conn:
        conn.execute("UPDATE invoices SET note=? WHERE digest=?", (note, digest))


def current_match_data() -> dict:
    with connect_db() as conn:
        return match_purchases(conn)


def folder_signature():
    items = []
    for p in BASE_DIR.glob("*.pdf"):
        try:
            st = p.stat()
            items.append((p.name.lower(), st.st_size, st.st_mtime_ns))
        except OSError:
            pass
    return tuple(sorted(items))


def export_excel(fields: list[str]) -> int:
    rows = all_active_rows()
    matches = current_match_data()
    match_map = matches["invoice_match_map"]
    wb = Workbook()
    ws = wb.active
    ws.title = "发票汇总"
    headers = fields + ["购买匹配", "人工确认", "备注"]
    ws.append(headers)
    for row in rows:
        values = [row[FIELD_DB_MAP[f]] for f in fields]
        info = match_map.get(str(row["digest"]))
        values.extend([
            f"{info['purchase_name']}（{info['kind']}{'，手动' if info.get('manual') else ''}）" if info else "",
            "是" if row["confirmed"] else "否",
            row["note"],
        ])
        ws.append(values)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    widths = {
        "文件名": 42, "发票号码": 24, "开票日期": 16, "购买方名称": 28,
        "销售方名称": 32, "项目名称": 46, "价税合计": 14, "金额（不含税）": 16,
        "税额": 12, "开票人": 12, "提取状态": 38, "购买匹配": 28,
        "人工确认": 12, "备注": 30,
    }
    for idx, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(idx)].width = widths.get(h, 18)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    pws = wb.create_sheet("购买记录")
    pws.append(["名称", "商品价格", "有快递费", "快递费", "商品发票", "快递费发票", "完整匹配"])
    for result in matches["purchase_results"]:
        p = result["purchase"]
        comp = {x["kind"]: x for x in result["components"]}
        item_inv = comp.get("商品", {}).get("invoice")
        ship_inv = comp.get("快递费", {}).get("invoice")
        pws.append([
            p["name"],
            p["item_price"],
            "是" if p["has_shipping"] else "否",
            p["shipping_fee"] if p["has_shipping"] else None,
            item_inv["filename"] if item_inv else "",
            ship_inv["filename"] if ship_inv else "",
            "是" if result["complete"] else "否",
        ])
    for cell in pws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for col, width in enumerate([28, 14, 12, 14, 42, 42, 12], 1):
        pws.column_dimensions[get_column_letter(col)].width = width
    wb.save(EXPORT_PATH)
    wb.close()
    return len(rows)


class InvoiceApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(f"{APP_TITLE} v{APP_VERSION}")
        self.geometry("1280x800")
        self.minsize(980, 640)
        init_db()
        self.status_var = tk.StringVar(value="就绪；自动监控发票文件夹已开启")
        self.count_var = tk.StringVar(value="0 张发票")
        self.search_var = tk.StringVar()
        self.search_scope_var = tk.StringVar(value="全部显示字段")
        self.only_unconfirmed_var = tk.BooleanVar(value=False)
        self.only_purchase_unmatched_var = tk.BooleanVar(value=False)
        self.editing_purchase_id = None
        self._sync_running = False
        self._cloud_backup_running = False
        self._cloud_backup_pending = False
        self._last_folder_signature = folder_signature()
        self._invoice_sort = (None, False)
        self._purchase_sort = (None, False)
        self._build_ui()
        self.refresh_all()
        self.after(150, lambda: self.run_background_sync(silent=True))
        self.after(FOLDER_POLL_MS, self._poll_folder)

    def _build_ui(self):
        top = ttk.Frame(self, padding=(12, 10, 12, 6))
        top.pack(fill="x")
        ttk.Label(top, text=f"当前目录：{BASE_DIR}").pack(side="left")
        ttk.Label(top, textvariable=self.count_var).pack(side="right")

        actions = ttk.Frame(self, padding=(12, 0, 12, 8))
        actions.pack(fill="x")
        ttk.Button(actions, text="从QQ邮箱获取京东发票", command=self.open_jd_fetch_dialog).pack(side="left")

        more_button = ttk.Menubutton(actions, text="更多")
        more_menu = tk.Menu(more_button, tearoff=False)
        more_menu.add_command(label="设置", command=self.open_settings)
        more_menu.add_separator()
        more_menu.add_command(label="刷新发票文件夹", command=self.run_background_sync)
        more_menu.add_command(label="导出 Excel", command=self.do_export)
        more_menu.add_command(label="数据备份与恢复", command=self.open_backup_dialog)
        more_menu.add_command(label="检查更新", command=self.check_for_updates)
        more_button["menu"] = more_menu
        more_button.pack(side="left", padx=(8, 0))

        self.notebook = ttk.Notebook(self)
        self.notebook.pack(fill="both", expand=True, padx=12, pady=(0, 8))
        self.invoice_tab = ttk.Frame(self.notebook)
        self.purchase_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.invoice_tab, text="发票管理")
        self.notebook.add(self.purchase_tab, text="购买记录")
        self._build_invoice_tab()
        self._build_purchase_tab()

        bottom = ttk.Frame(self, padding=(12, 0, 12, 10))
        bottom.pack(fill="x")
        ttk.Label(bottom, textvariable=self.status_var).pack(side="right")

    def set_status(self, text):
        self.after(0, lambda: self.status_var.set(text))

    def _poll_folder(self):
        try:
            sig = folder_signature()
            if sig != self._last_folder_signature and not self._sync_running:
                self._last_folder_signature = sig
                self.run_background_sync(silent=True, source="检测到文件夹变化")
        finally:
            self.after(FOLDER_POLL_MS, self._poll_folder)

    def _sort_tree(self, tree, column, numeric_columns, state_attr):
        current_col, desc = getattr(self, state_attr)
        desc = (not desc) if current_col == column else False
        setattr(self, state_attr, (column, desc))

        def key(item_id):
            value = tree.set(item_id, column)
            if column in numeric_columns:
                try:
                    return (0, float(str(value).replace("¥", "").strip()))
                except Exception:
                    return (1, float("inf"))
            return (0, str(value).lower())

        items = list(tree.get_children(""))
        items.sort(key=key, reverse=desc)
        for idx, item_id in enumerate(items):
            tree.move(item_id, "", idx)

    def _visible_search_scopes(self):
        return ["全部显示字段"] + selected_fields() + ["购买匹配", "人工确认", "备注"]

    def _build_invoice_tab(self):
        filters = ttk.Frame(self.invoice_tab, padding=(8, 8, 8, 6))
        filters.pack(fill="x")
        ttk.Label(filters, text="搜索：").pack(side="left")
        self.search_scope_box = ttk.Combobox(
            filters,
            textvariable=self.search_scope_var,
            values=self._visible_search_scopes(),
            state="readonly",
            width=15,
        )
        self.search_scope_box.pack(side="left", padx=(4, 4))
        self.search_scope_box.bind("<<ComboboxSelected>>", lambda _e: self.refresh_invoice_table())
        ttk.Entry(filters, textvariable=self.search_var, width=30).pack(side="left", padx=(0, 12))
        self.search_var.trace_add("write", lambda *_: self.refresh_invoice_table())
        ttk.Checkbutton(
            filters,
            text="只看未人工确认",
            variable=self.only_unconfirmed_var,
            command=self.refresh_invoice_table,
        ).pack(side="left", padx=(0, 12))
        ttk.Checkbutton(
            filters,
            text="只看未匹配购买记录的发票",
            variable=self.only_purchase_unmatched_var,
            command=self.refresh_invoice_table,
        ).pack(side="left")

        table_frame = ttk.Frame(self.invoice_tab, padding=(8, 0, 8, 6))
        table_frame.pack(fill="both", expand=True)
        self.invoice_tree = ttk.Treeview(table_frame, show="headings", selectmode="extended")
        ybar = ttk.Scrollbar(table_frame, orient="vertical", command=self.invoice_tree.yview)
        xbar = ttk.Scrollbar(table_frame, orient="horizontal", command=self.invoice_tree.xview)
        self.invoice_tree.configure(yscrollcommand=ybar.set, xscrollcommand=xbar.set)
        self.invoice_tree.grid(row=0, column=0, sticky="nsew")
        ybar.grid(row=0, column=1, sticky="ns")
        xbar.grid(row=1, column=0, sticky="ew")
        table_frame.rowconfigure(0, weight=1)
        table_frame.columnconfigure(0, weight=1)
        self.invoice_tree.bind("<Double-1>", self.on_invoice_double_click)

        actions = ttk.Frame(self.invoice_tab, padding=(8, 0, 8, 8))
        actions.pack(fill="x")
        ttk.Button(actions, text="人工确认标记", command=self.mark_selected_confirmed).pack(side="left")
        ttk.Button(actions, text="编辑备注", command=self.edit_selected_note).pack(side="left", padx=(6, 0))
        ttk.Button(actions, text="删除选中发票", command=self.delete_selected_invoices).pack(side="left", padx=(6, 0))
        ttk.Label(actions, text="提示：Ctrl/Shift 可多选；点击表头可排序").pack(side="right")

    def _searchable_values(self, row, match_info):
        def s(v):
            if v is None:
                return ""
            if isinstance(v, float):
                return f"{v:.2f}"
            return str(v)

        data = {field: s(row[db_key]) for field, db_key in FIELD_DB_MAP.items()}
        data["购买匹配"] = (f"{match_info['purchase_name']} {match_info['kind']}" + (" 手动" if match_info.get("manual") else "")) if match_info else ""
        data["人工确认"] = "已确认 是 yes ✓" if row["confirmed"] else "未确认 否 no"
        data["备注"] = s(row["note"])
        return data

    def refresh_invoice_table(self):
        if not hasattr(self, "invoice_tree"):
            return
        fields = selected_fields()
        scopes = self._visible_search_scopes()
        self.search_scope_box["values"] = scopes
        if self.search_scope_var.get() not in scopes:
            self.search_scope_var.set("全部显示字段")
        columns = ["_digest"] + fields + ["购买匹配", "人工确认", "备注"]
        self.invoice_tree["columns"] = columns
        self.invoice_tree.column("_digest", width=0, stretch=False)
        self.invoice_tree.heading("_digest", text="")
        widths = {
            "文件名": 280, "项目名称": 300, "价税合计": 100, "发票号码": 180,
            "开票日期": 120, "购买方名称": 220, "销售方名称": 240,
            "金额（不含税）": 120, "税额": 90, "开票人": 100, "提取状态": 240,
            "购买匹配": 190, "人工确认": 90, "备注": 220,
        }
        numeric = {"价税合计", "金额（不含税）", "税额"}
        for field in fields + ["购买匹配", "人工确认", "备注"]:
            self.invoice_tree.heading(
                field,
                text=field,
                command=lambda c=field: self._sort_tree(self.invoice_tree, c, numeric, "_invoice_sort"),
            )
            self.invoice_tree.column(field, width=widths.get(field, 150), minwidth=70)
        for item in self.invoice_tree.get_children():
            self.invoice_tree.delete(item)

        match_map = current_match_data()["invoice_match_map"]
        rows = all_active_rows()
        query = self.search_var.get().strip().lower()
        scope = self.search_scope_var.get()
        filtered = []
        for row in rows:
            digest = str(row["digest"])
            info = match_map.get(digest)
            if self.only_unconfirmed_var.get() and row["confirmed"]:
                continue
            if self.only_purchase_unmatched_var.get() and info:
                continue
            if query:
                values = self._searchable_values(row, info)
                if scope == "全部显示字段":
                    visible = [name for name in scopes if name != "全部显示字段"]
                    haystack = " ".join(values.get(name, "") for name in visible)
                else:
                    haystack = values.get(scope, "")
                if query not in haystack.lower():
                    continue
            filtered.append((row, info))

        for row, info in filtered:
            values = [row["digest"]]
            for field in fields:
                value = row[FIELD_DB_MAP[field]]
                if field in numeric and isinstance(value, (int, float)):
                    value = f"{value:.2f}"
                values.append("" if value is None else value)
            values.extend([
                f"✓ {info['purchase_name']}（{info['kind']}{'，手动' if info.get('manual') else ''}）" if info else "",
                "✓" if row["confirmed"] else "",
                row["note"],
            ])
            self.invoice_tree.insert("", "end", values=values)
        self.count_var.set(f"{len(rows)} 张发票")

    def selected_digests(self):
        result = []
        for item in self.invoice_tree.selection():
            values = self.invoice_tree.item(item, "values")
            if values:
                result.append(str(values[0]))
        return result

    def mark_selected_confirmed(self):
        digests = self.selected_digests()
        if not digests:
            messagebox.showinfo(APP_TITLE, "请先选中至少一张发票。")
            return
        with connect_db() as conn:
            placeholders = ",".join("?" for _ in digests)
            rows = conn.execute(
                f"SELECT digest, confirmed FROM invoices WHERE digest IN ({placeholders})",
                digests,
            ).fetchall()
        all_confirmed = bool(rows) and all(bool(r["confirmed"]) for r in rows)
        set_confirmed_many(digests, not all_confirmed)
        self.refresh_invoice_table()
        self.schedule_auto_cloud_backup()

    def on_invoice_double_click(self, event):
        if self.invoice_tree.identify_region(event.x, event.y) != "cell":
            return
        col = self.invoice_tree.identify_column(event.x)
        try:
            idx = int(col.lstrip("#")) - 1
            name = self.invoice_tree["columns"][idx]
        except Exception:
            return
        if name == "人工确认":
            self.mark_selected_confirmed()

    def edit_selected_note(self):
        digests = self.selected_digests()
        if len(digests) != 1:
            messagebox.showinfo(APP_TITLE, "编辑备注时请只选中一张发票。")
            return
        digest = digests[0]
        with connect_db() as conn:
            row = conn.execute("SELECT filename, note FROM invoices WHERE digest=?", (digest,)).fetchone()
        if not row:
            return
        win = tk.Toplevel(self)
        win.title("编辑备注")
        win.transient(self)
        win.grab_set()
        win.geometry("520x230")
        ttk.Label(win, text=row["filename"], padding=(12, 12, 12, 6)).pack(anchor="w")
        text = tk.Text(win, height=6, wrap="word")
        text.pack(fill="both", expand=True, padx=12, pady=(0, 8))
        text.insert("1.0", row["note"] or "")
        text.focus_set()
        buttons = ttk.Frame(win, padding=(12, 0, 12, 12))
        buttons.pack(fill="x")

        def save():
            set_note(digest, text.get("1.0", "end").strip())
            win.destroy()
            self.refresh_invoice_table()
            self.schedule_auto_cloud_backup()

        ttk.Button(buttons, text="保存", command=save).pack(side="right")
        ttk.Button(buttons, text="取消", command=win.destroy).pack(side="right", padx=(0, 6))

    def delete_selected_invoices(self):
        digests = self.selected_digests()
        if not digests:
            messagebox.showinfo(APP_TITLE, "请先选中要删除的发票。")
            return
        with connect_db() as conn:
            placeholders = ",".join("?" for _ in digests)
            rows = conn.execute(
                f"SELECT digest, filename FROM invoices WHERE digest IN ({placeholders})",
                digests,
            ).fetchall()
        names = [str(r["filename"]) for r in rows]
        preview = "\n".join(names[:8]) + ("\n……" if len(names) > 8 else "")
        if not messagebox.askyesno(
            APP_TITLE,
            f"确定永久删除选中的 {len(names)} 个 PDF 文件？\n\n"
            f"文件会从当前发票文件夹中删除，无法在程序内撤销。\n\n{preview}",
        ):
            return
        errors = []
        deleted = 0
        for name in names:
            path = (BASE_DIR / name).resolve()
            try:
                if path.parent != BASE_DIR.resolve():
                    raise RuntimeError("文件路径不在当前发票目录")
                if path.exists():
                    path.unlink()
                deleted += 1
            except Exception as e:
                errors.append(f"{name}: {type(e).__name__}: {e}")
        self._last_folder_signature = folder_signature()
        self.run_background_sync(silent=True, source="删除发票后刷新")
        if errors:
            messagebox.showwarning(
                APP_TITLE,
                f"成功删除 {deleted} 个，失败 {len(errors)} 个：\n\n" + "\n".join(errors[:6]),
            )

    def _build_purchase_tab(self):
        form = ttk.LabelFrame(self.purchase_tab, text="新增 / 编辑购买记录", padding=10)
        form.pack(fill="x", padx=8, pady=(8, 6))
        ttk.Label(form, text="名称").grid(row=0, column=0, sticky="w")
        self.purchase_name_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.purchase_name_var, width=28).grid(row=1, column=0, padx=(0, 12), sticky="ew")
        ttk.Label(form, text="商品价格").grid(row=0, column=1, sticky="w")
        self.purchase_price_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.purchase_price_var, width=14).grid(row=1, column=1, padx=(0, 12), sticky="ew")
        self.has_shipping_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(
            form,
            text="有快递费",
            variable=self.has_shipping_var,
            command=self._update_shipping_entry_state,
        ).grid(row=1, column=2, padx=(0, 12), sticky="w")
        ttk.Label(form, text="快递费").grid(row=0, column=3, sticky="w")
        self.shipping_fee_var = tk.StringVar(value="")
        self.shipping_entry = ttk.Entry(form, textvariable=self.shipping_fee_var, width=14)
        self.shipping_entry.grid(row=1, column=3, padx=(0, 12), sticky="ew")
        self._update_shipping_entry_state()
        self.purchase_save_button = ttk.Button(form, text="新增", command=self.save_purchase_form)
        self.purchase_save_button.grid(row=1, column=4, padx=(4, 4))
        ttk.Button(form, text="取消编辑", command=self.reset_purchase_form).grid(row=1, column=5, padx=(4, 0))

        table_frame = ttk.Frame(self.purchase_tab, padding=(8, 0, 8, 6))
        table_frame.pack(fill="both", expand=True)
        columns = ("id", "名称", "商品价格", "有快递费", "快递费", "商品发票", "快递费发票", "状态")
        self.purchase_tree = ttk.Treeview(
            table_frame,
            columns=columns,
            show="headings",
            selectmode="extended",
            height=11,
        )
        p_ybar = ttk.Scrollbar(table_frame, orient="vertical", command=self.purchase_tree.yview)
        self.purchase_tree.configure(yscrollcommand=p_ybar.set)
        widths = [60, 220, 110, 90, 100, 280, 280, 100]
        numeric = {"id", "商品价格", "快递费"}
        for name, width in zip(columns, widths):
            self.purchase_tree.heading(
                name,
                text=name,
                command=lambda c=name: self._sort_tree(self.purchase_tree, c, numeric, "_purchase_sort"),
            )
            self.purchase_tree.column(
                name,
                width=width,
                minwidth=60,
                stretch=name in {"名称", "商品发票", "快递费发票"},
            )
        self.purchase_tree.grid(row=0, column=0, sticky="nsew")
        p_ybar.grid(row=0, column=1, sticky="ns")
        table_frame.rowconfigure(0, weight=1)
        table_frame.columnconfigure(0, weight=1)
        self.purchase_tree.bind("<Double-1>", lambda _e: self.load_selected_purchase_for_edit())

        p_actions = ttk.Frame(self.purchase_tab, padding=(8, 0, 8, 6))
        p_actions.pack(fill="x")
        ttk.Button(p_actions, text="编辑选中项", command=self.load_selected_purchase_for_edit).pack(side="left")
        ttk.Button(p_actions, text="调整发票关联", command=self.open_manual_match_dialog).pack(side="left", padx=(6, 0))
        ttk.Button(p_actions, text="删除选中项", command=self.delete_selected_purchase).pack(side="left", padx=(6, 0))
        ttk.Button(p_actions, text="一键清除全部", command=self.clear_all_purchases).pack(side="left", padx=(6, 0))
        self.purchase_summary_var = tk.StringVar(value="")
        ttk.Label(p_actions, textvariable=self.purchase_summary_var).pack(side="right")

        compare = ttk.Frame(self.purchase_tab, padding=(8, 0, 8, 8))
        compare.pack(fill="both", expand=True)
        compare.columnconfigure(0, weight=1)
        compare.columnconfigure(1, weight=1)
        compare.rowconfigure(1, weight=1)
        ttk.Label(compare, text="已记录购买，但缺少对应发票").grid(row=0, column=0, sticky="w", padx=(0, 8))
        ttk.Label(compare, text="有发票，但没有对应购买记录").grid(row=0, column=1, sticky="w", padx=(8, 0))
        self.missing_tree = ttk.Treeview(compare, columns=("购买", "类型", "金额"), show="headings", height=6)
        for c, w in [("购买", 220), ("类型", 90), ("金额", 100)]:
            self.missing_tree.heading(
                c,
                text=c,
                command=lambda col=c: self._sort_tree(self.missing_tree, col, {"金额"}, "_purchase_sort"),
            )
            self.missing_tree.column(c, width=w)
        self.missing_tree.grid(row=1, column=0, sticky="nsew", padx=(0, 8), pady=(4, 0))
        self.unused_invoice_tree = ttk.Treeview(
            compare,
            columns=("文件名", "价税合计"),
            show="headings",
            height=6,
        )
        self.unused_invoice_tree.heading(
            "文件名",
            text="文件名",
            command=lambda: self._sort_tree(self.unused_invoice_tree, "文件名", set(), "_purchase_sort"),
        )
        self.unused_invoice_tree.heading(
            "价税合计",
            text="价税合计",
            command=lambda: self._sort_tree(self.unused_invoice_tree, "价税合计", {"价税合计"}, "_purchase_sort"),
        )
        self.unused_invoice_tree.column("文件名", width=360)
        self.unused_invoice_tree.column("价税合计", width=100)
        self.unused_invoice_tree.grid(row=1, column=1, sticky="nsew", padx=(8, 0), pady=(4, 0))

    def open_manual_match_dialog(self):
        ids = self.selected_purchase_ids()
        if len(ids) != 1:
            messagebox.showinfo(APP_TITLE, "调整关联时请只选中一条购买记录。")
            return
        purchase_id = ids[0]
        with connect_db() as conn:
            purchase = get_purchase(conn, purchase_id)
        if not purchase:
            return

        components = required_components(purchase)
        component_by_kind = {x["kind"]: x for x in components}

        win = tk.Toplevel(self)
        win.title("调整发票关联")
        win.transient(self)
        win.grab_set()
        win.geometry("860x560")
        outer = ttk.Frame(win, padding=14)
        outer.pack(fill="both", expand=True)

        ttk.Label(
            outer,
            text=f"购买记录 #{purchase_id}：{purchase['name']}",
            font=("TkDefaultFont", 10, "bold"),
        ).pack(anchor="w")
        ttk.Label(
            outer,
            text="手动关联优先于自动金额匹配；恢复自动匹配后，程序会重新按价格分配。",
        ).pack(anchor="w", pady=(4, 10))

        controls = ttk.Frame(outer)
        controls.pack(fill="x", pady=(0, 8))
        ttk.Label(controls, text="要调整：").pack(side="left")
        kind_var = tk.StringVar(value=components[0]["kind"])
        kind_box = ttk.Combobox(
            controls,
            textvariable=kind_var,
            values=[x["kind"] for x in components],
            state="readonly",
            width=10,
        )
        kind_box.pack(side="left", padx=(4, 14))
        ttk.Label(controls, text="搜索发票：").pack(side="left")
        search_var = tk.StringVar()
        ttk.Entry(controls, textvariable=search_var, width=32).pack(side="left", padx=(4, 0))

        current_var = tk.StringVar()
        ttk.Label(outer, textvariable=current_var).pack(anchor="w", pady=(0, 6))

        frame = ttk.Frame(outer)
        frame.pack(fill="both", expand=True)
        columns = ("_digest", "文件名", "价税合计", "项目名称", "当前关联")
        tree = ttk.Treeview(frame, columns=columns, show="headings", selectmode="browse")
        tree.heading("_digest", text="")
        tree.column("_digest", width=0, stretch=False)
        for name, width in [("文件名", 320), ("价税合计", 100), ("项目名称", 220), ("当前关联", 240)]:
            tree.heading(name, text=name)
            tree.column(name, width=width, minwidth=80)
        ybar = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=ybar.set)
        tree.grid(row=0, column=0, sticky="nsew")
        ybar.grid(row=0, column=1, sticky="ns")
        frame.rowconfigure(0, weight=1)
        frame.columnconfigure(0, weight=1)

        def fresh_data():
            with connect_db() as conn:
                current_purchase = get_purchase(conn, purchase_id)
                data = match_purchases(conn)
                invoices = conn.execute(
                    "SELECT digest, filename, total, project FROM invoices WHERE active=1 ORDER BY filename COLLATE NOCASE"
                ).fetchall()
            result = next(
                (x for x in data["purchase_results"] if int(x["purchase"]["id"]) == purchase_id),
                None,
            )
            return current_purchase, data, invoices, result

        def refresh_dialog(*_args):
            for item in tree.get_children():
                tree.delete(item)
            current_purchase, data, invoices, result = fresh_data()
            if not current_purchase or not result:
                current_var.set("购买记录已不存在。")
                return
            kind = kind_var.get()
            component = next((x for x in result["components"] if x["kind"] == kind), None)
            current_digest = None
            if component and component.get("invoice") is not None:
                current_digest = str(component["invoice"]["digest"])
                mode = "手动关联" if component.get("manual") else "自动匹配"
                current_var.set(f"当前：{mode} → {component['invoice']['filename']}")
            elif component and component.get("manual_missing"):
                current_var.set("当前：手动关联的发票已不在当前文件夹；恢复该 PDF 后会重新生效。")
            else:
                current_var.set("当前：未匹配")

            query = search_var.get().strip().lower()
            match_map = data["invoice_match_map"]
            selected_item = None
            for inv in invoices:
                total = inv["total"]
                total_text = f"{total:.2f}" if isinstance(total, (int, float)) else ""
                project = str(inv["project"] or "")
                haystack = f"{inv['filename']} {total_text} {project}".lower()
                if query and query not in haystack:
                    continue
                digest = str(inv["digest"])
                info = match_map.get(digest)
                association = ""
                if info:
                    mode = "手动" if info.get("manual") else "自动"
                    association = f"{info['purchase_name']}（{info['kind']}，{mode}）"
                item_id = tree.insert(
                    "",
                    "end",
                    values=(digest, inv["filename"], total_text, project, association),
                )
                if digest == current_digest:
                    selected_item = item_id
            if selected_item:
                tree.selection_set(selected_item)
                tree.see(selected_item)

        def set_selected_manual():
            selection = tree.selection()
            if len(selection) != 1:
                messagebox.showinfo(APP_TITLE, "请先选择一张发票。", parent=win)
                return
            values = tree.item(selection[0], "values")
            digest = str(values[0])
            filename = str(values[1])
            kind = kind_var.get()
            expected = component_by_kind[kind]["price"]
            try:
                actual = float(values[2])
            except Exception:
                actual = None
            if actual is not None and int(round(actual * 100)) != int(round(expected * 100)):
                if not messagebox.askyesno(
                    APP_TITLE,
                    f"金额不同，仍要手动关联吗？

购买记录“{kind}”：¥{expected:.2f}
"
                    f"所选发票：¥{actual:.2f}
{filename}",
                    parent=win,
                ):
                    return
            if not self._create_local_safety_backup_or_block("before-manual-invoice-match"):
                return
            try:
                with connect_db() as conn:
                    set_manual_match(conn, purchase_id, kind, digest)
            except ValueError as e:
                messagebox.showwarning(APP_TITLE, str(e), parent=win)
                return
            self.refresh_all()
            self.schedule_auto_cloud_backup()
            self.set_status(f"已手动关联：{purchase['name']}（{kind}） → {filename}")
            refresh_dialog()

        def restore_auto():
            kind = kind_var.get()
            if not self._create_local_safety_backup_or_block("before-clear-manual-invoice-match"):
                return
            with connect_db() as conn:
                clear_manual_match(conn, purchase_id, kind)
            self.refresh_all()
            self.schedule_auto_cloud_backup()
            self.set_status(f"已恢复自动匹配：{purchase['name']}（{kind}）")
            refresh_dialog()

        kind_box.bind("<<ComboboxSelected>>", refresh_dialog)
        search_var.trace_add("write", refresh_dialog)

        buttons = ttk.Frame(outer)
        buttons.pack(fill="x", pady=(10, 0))
        ttk.Button(buttons, text="设为手动关联", command=set_selected_manual).pack(side="left")
        ttk.Button(buttons, text="恢复自动匹配", command=restore_auto).pack(side="left", padx=(8, 0))
        ttk.Button(buttons, text="关闭", command=win.destroy).pack(side="right")
        refresh_dialog()

    def _update_shipping_entry_state(self):
        if not hasattr(self, "shipping_entry"):
            return
        if self.has_shipping_var.get():
            self.shipping_entry.configure(state="normal")
        else:
            self.shipping_entry.configure(state="disabled")
            self.shipping_fee_var.set("")

    def reset_purchase_form(self):
        self.editing_purchase_id = None
        self.purchase_name_var.set("")
        self.purchase_price_var.set("")
        self.has_shipping_var.set(False)
        self.shipping_fee_var.set("")
        self._update_shipping_entry_state()
        self.purchase_save_button.configure(text="新增")

    def _create_local_safety_backup_or_block(self, reason: str) -> bool:
        try:
            path = create_local_safety_backup(DB_PATH, LOCAL_BACKUP_DIR, reason=reason)
            hide_windows_file(LOCAL_BACKUP_DIR)
            self.set_status(f"已创建本地安全备份：{path.name}")
            return True
        except Exception as e:
            messagebox.showerror(
                APP_TITLE,
                f"安全备份失败。为避免数据丢失，本次操作已取消。\n\n{type(e).__name__}: {e}",
            )
            return False

    def save_purchase_form(self):
        try:
            name = self.purchase_name_var.get().strip()
            item_price = self.purchase_price_var.get().strip()
            has_shipping = self.has_shipping_var.get()
            shipping_fee = self.shipping_fee_var.get().strip() if has_shipping else 0
            if self.editing_purchase_id is not None:
                if not self._create_local_safety_backup_or_block("before-purchase-edit"):
                    return
            with connect_db() as conn:
                if self.editing_purchase_id is None:
                    add_purchase(conn, name, item_price, has_shipping, shipping_fee)
                else:
                    update_purchase(conn, self.editing_purchase_id, name, item_price, has_shipping, shipping_fee)
            self.reset_purchase_form()
            self.refresh_all()
            self.schedule_auto_cloud_backup()
        except ValueError as e:
            messagebox.showwarning(APP_TITLE, str(e))
        except Exception as e:
            messagebox.showerror(APP_TITLE, f"保存购买记录失败：{type(e).__name__}: {e}")

    def selected_purchase_ids(self):
        ids = []
        for sel in self.purchase_tree.selection():
            values = self.purchase_tree.item(sel, "values")
            if values:
                try:
                    ids.append(int(values[0]))
                except Exception:
                    pass
        return ids

    def load_selected_purchase_for_edit(self):
        ids = self.selected_purchase_ids()
        if len(ids) != 1:
            messagebox.showinfo(APP_TITLE, "编辑时请只选中一条购买记录。")
            return
        purchase_id = ids[0]
        with connect_db() as conn:
            p = get_purchase(conn, purchase_id)
        if not p:
            return
        self.editing_purchase_id = purchase_id
        self.purchase_name_var.set(p["name"])
        self.purchase_price_var.set(f"{p['item_price']:.2f}")
        self.has_shipping_var.set(bool(p["has_shipping"]))
        self.shipping_fee_var.set(f"{p['shipping_fee']:.2f}" if p["has_shipping"] else "")
        self._update_shipping_entry_state()
        self.purchase_save_button.configure(text="保存修改")

    def delete_selected_purchase(self):
        ids = self.selected_purchase_ids()
        if not ids:
            messagebox.showinfo(APP_TITLE, "请先选中要删除的购买记录。")
            return
        if not messagebox.askyesno(APP_TITLE, f"确定删除选中的 {len(ids)} 条购买记录？"):
            return
        if not self._create_local_safety_backup_or_block("before-purchase-delete"):
            return
        with connect_db() as conn:
            for purchase_id in ids:
                delete_purchase(conn, purchase_id)
        if self.editing_purchase_id in ids:
            self.reset_purchase_form()
        self.refresh_all()
        self.schedule_auto_cloud_backup()

    def clear_all_purchases(self):
        with connect_db() as conn:
            count = conn.execute("SELECT COUNT(*) FROM purchases").fetchone()[0]
        if not count:
            messagebox.showinfo(APP_TITLE, "当前没有购买记录。")
            return
        if not messagebox.askyesno(
            APP_TITLE,
            f"确定清空全部 {count} 条购买记录？\n\n此操作不会删除任何 PDF 发票。",
        ):
            return
        if not self._create_local_safety_backup_or_block("before-purchase-clear-all"):
            return
        with connect_db() as conn:
            clear_purchases(conn)
        self.reset_purchase_form()
        self.refresh_all()
        self.schedule_auto_cloud_backup()

    def refresh_purchase_tab(self):
        if not hasattr(self, "purchase_tree"):
            return
        with connect_db() as conn:
            data = match_purchases(conn)
        for tree in (self.purchase_tree, self.missing_tree, self.unused_invoice_tree):
            for item in tree.get_children():
                tree.delete(item)

        total_cents = 0
        for result in data["purchase_results"]:
            p = result["purchase"]
            total_cents += int(round(float(p["item_price"]) * 100))
            if p["has_shipping"]:
                total_cents += int(round(float(p["shipping_fee"]) * 100))
            comp = {x["kind"]: x for x in result["components"]}
            item_inv = comp.get("商品", {}).get("invoice")
            ship_inv = comp.get("快递费", {}).get("invoice")
            item_text = item_inv["filename"] if item_inv else ""
            ship_text = ship_inv["filename"] if ship_inv else ""
            if item_inv and comp.get("商品", {}).get("manual"):
                item_text = "手动：" + item_text
            if ship_inv and comp.get("快递费", {}).get("manual"):
                ship_text = "手动：" + ship_text
            has_manual = any(x.get("manual") for x in result["components"])
            status_text = "✓ 完整" if result["complete"] else "缺发票"
            if has_manual:
                status_text += " · 手动"
            self.purchase_tree.insert("", "end", values=(
                p["id"],
                p["name"],
                f"{p['item_price']:.2f}",
                "✓" if p["has_shipping"] else "",
                f"{p['shipping_fee']:.2f}" if p["has_shipping"] else "",
                item_text,
                ship_text,
                status_text,
            ))
        for item in data["missing_components"]:
            self.missing_tree.insert(
                "",
                "end",
                values=(item["purchase_name"], item["kind"], f"{item['price']:.2f}"),
            )
        for inv in data["unused_invoices"]:
            total = inv["total"]
            self.unused_invoice_tree.insert(
                "",
                "end",
                values=(inv["filename"], f"{total:.2f}" if isinstance(total, (int, float)) else ""),
            )
        self.purchase_summary_var.set(
            f"购买 {len(data['purchase_results'])} 项 | 总金额 ¥{total_cents / 100:.2f} | "
            f"缺票 {len(data['missing_components'])} 项 | 未对应发票 {len(data['unused_invoices'])} 张"
        )

    def refresh_all(self):
        self.refresh_invoice_table()
        self.refresh_purchase_tab()

    def run_background_sync(self, silent=False, source="手动刷新"):
        if self._sync_running:
            return
        self._sync_running = True

        def worker():
            try:
                self.set_status(f"{source}：正在扫描发票文件夹...")
                stats = sync_folder(progress=self.set_status)
                self._last_folder_signature = folder_signature()
                self.after(0, self.refresh_all)
                msg = (
                    f"刷新完成：{stats['total']} 张；复用 {stats['reused']}，重新解析 {stats['parsed']}，"
                    f"移除 {stats['removed']}，异常 {stats['failed']}"
                )
                self.set_status(msg)
                if not silent and stats["failed"]:
                    self.after(
                        0,
                        lambda: messagebox.showwarning(
                            APP_TITLE,
                            msg + "\n\n有解析异常，可显示“提取状态”字段查看。",
                        ),
                    )
            except Exception as e:
                detail = traceback.format_exc()
                self.set_status("刷新失败")
                self.after(
                    0,
                    lambda: messagebox.showerror(
                        APP_TITLE,
                        f"刷新失败：{type(e).__name__}: {e}\n\n" + detail[-1800:],
                    ),
                )
            finally:
                self._sync_running = False

        threading.Thread(target=worker, daemon=True).start()

    def open_settings(self):
        win = tk.Toplevel(self)
        win.title("设置")
        win.transient(self)
        win.grab_set()
        win.geometry("620x650")
        outer = ttk.Frame(win, padding=14)
        outer.pack(fill="both", expand=True)

        ttk.Label(outer, text="QQ邮箱（京东发票获取）").pack(anchor="w")
        email_var = tk.StringVar(value=get_setting("qq_email", ""))
        ttk.Label(outer, text="QQ邮箱").pack(anchor="w", pady=(10, 2))
        ttk.Entry(outer, textvariable=email_var).pack(fill="x")
        ttk.Label(outer, text="QQ邮箱16位授权码").pack(anchor="w", pady=(10, 2))
        auth_var = tk.StringVar()
        ttk.Entry(outer, textvariable=auth_var, show="•").pack(fill="x")
        has_saved_auth = bool(get_setting("qq_auth_code", ""))
        auth_tip = "已保存授权码；此处留空将继续使用原授权码。" if has_saved_auth else "尚未保存授权码。"
        ttk.Label(outer, text=auth_tip + " Windows 下使用 DPAPI 加密保存。").pack(anchor="w", pady=(3, 8))

        ttk.Separator(outer).pack(fill="x", pady=8)
        ttk.Label(outer, text="发票表显示字段").pack(anchor="w", pady=(4, 6))
        current = set(selected_fields())
        field_vars = {}
        fields_box = ttk.Frame(outer)
        fields_box.pack(fill="both", expand=True)
        for field in AVAILABLE_FIELDS:
            var = tk.BooleanVar(value=field in current)
            field_vars[field] = var
            ttk.Checkbutton(fields_box, text=field, variable=var).pack(anchor="w")

        buttons = ttk.Frame(outer)
        buttons.pack(fill="x", pady=(10, 0))

        def save():
            fields = [f for f, v in field_vars.items() if v.get()]
            if not fields:
                messagebox.showwarning(APP_TITLE, "至少选择一个显示字段。", parent=win)
                return
            email_addr = email_var.get().strip()
            if email_addr and "@" not in email_addr:
                messagebox.showwarning(APP_TITLE, "QQ邮箱格式看起来不正确。", parent=win)
                return
            try:
                with connect_db() as conn:
                    set_setting_conn(conn, "qq_email", email_addr)
                    set_setting_conn(conn, "selected_fields", json.dumps(fields, ensure_ascii=False))
                    if auth_var.get().strip():
                        set_setting_conn(conn, "qq_auth_code", protect_secret(auth_var.get().strip()))
            except Exception as e:
                messagebox.showerror(APP_TITLE, f"保存设置失败：{type(e).__name__}: {e}", parent=win)
                return
            win.destroy()
            self.refresh_all()
            self.set_status("设置已保存")
            self.schedule_auto_cloud_backup()

        def clear_auth():
            if messagebox.askyesno(APP_TITLE, "确定清除已保存的 QQ 邮箱授权码？", parent=win):
                set_setting("qq_auth_code", "")
                auth_var.set("")

        ttk.Button(buttons, text="保存", command=save).pack(side="right")
        ttk.Button(buttons, text="取消", command=win.destroy).pack(side="right", padx=(0, 6))
        ttk.Button(buttons, text="清除授权码", command=clear_auth).pack(side="left")

    def open_jd_fetch_dialog(self):
        email_addr = get_setting("qq_email", "").strip()
        enc = get_setting("qq_auth_code", "").strip()
        if not email_addr or not enc:
            messagebox.showinfo(APP_TITLE, "请先在“设置”里填写 QQ 邮箱和16位授权码。")
            self.open_settings()
            return
        try:
            auth_code = unprotect_secret(enc)
        except Exception as e:
            messagebox.showerror(
                APP_TITLE,
                f"读取已保存授权码失败：{type(e).__name__}: {e}\n请在设置里重新保存授权码。",
            )
            return
        if not auth_code:
            messagebox.showinfo(APP_TITLE, "请在“设置”里重新保存 QQ 邮箱授权码。")
            return

        win = tk.Toplevel(self)
        win.title("从QQ邮箱获取京东发票")
        win.transient(self)
        win.grab_set()
        win.geometry("560x360")
        outer = ttk.Frame(win, padding=14)
        outer.pack(fill="both", expand=True)
        ttk.Label(outer, text=f"邮箱：{email_addr}").pack(anchor="w")
        last = get_setting("jd_last_fetch", "")
        last_text = last.replace("T", " ")[:16] if last else "尚未成功获取过"
        ttk.Label(outer, text=f"上次成功获取：{last_text}").pack(anchor="w", pady=(4, 12))
        mode_var = tk.StringVar(value="last" if last else "all")
        ttk.Radiobutton(
            outer,
            text="从上次成功获取之后（推荐）",
            variable=mode_var,
            value="last",
            state="normal" if last else "disabled",
        ).pack(anchor="w", pady=3)
        ttk.Radiobutton(outer, text="自定义起始时间", variable=mode_var, value="custom").pack(anchor="w", pady=3)
        custom_var = tk.StringVar(value=datetime.now().strftime("%Y-%m-%d 00:00"))
        ttk.Entry(outer, textvariable=custom_var).pack(fill="x", padx=(24, 0), pady=(2, 6))
        ttk.Label(
            outer,
            text="格式：YYYY-MM-DD 或 YYYY-MM-DD HH:MM；只填日期默认 00:00",
        ).pack(anchor="w", padx=(24, 0), pady=(0, 6))
        ttk.Radiobutton(outer, text="扫描全部历史收件箱", variable=mode_var, value="all").pack(anchor="w", pady=3)
        buttons = ttk.Frame(outer)
        buttons.pack(fill="x", pady=(18, 0))

        def start():
            mode = mode_var.get()
            start_time = last if mode == "last" else custom_var.get().strip() if mode == "custom" else None
            win.destroy()
            self.run_jd_fetch(email_addr, auth_code, start_time)

        ttk.Button(buttons, text="开始获取", command=start).pack(side="right")
        ttk.Button(buttons, text="取消", command=win.destroy).pack(side="right", padx=(0, 6))

    def run_jd_fetch(self, email_addr, auth_code, start_time):
        def worker():
            started = datetime.now()
            try:
                self.set_status("正在连接 QQ 邮箱...")
                result = fetch_jd_invoices(
                    email_addr=email_addr,
                    auth_code=auth_code,
                    output_dir=BASE_DIR,
                    start_time=start_time,
                    progress=self.set_status,
                )
                if result["failed"] == 0:
                    set_setting("jd_last_fetch", started.isoformat(timespec="seconds"))
                stats = sync_folder(progress=self.set_status)
                self._last_folder_signature = folder_signature()
                self.after(0, self.refresh_all)
                msg = (
                    f"邮箱扫描 {result['emails']} 封，发现 {result['links']} 个 PDF 链接。\n"
                    f"新下载 {result['downloaded']}，已存在 {result['skipped']}，失败 {result['failed']}。\n"
                    f"当前文件夹共 {stats['total']} 张 PDF。"
                )
                self.set_status(f"邮箱获取完成：新下载 {result['downloaded']}，失败 {result['failed']}")
                if result["failed"]:
                    detail = "\n".join(result["errors"][:5])
                    detail += "\n……" if len(result["errors"]) > 5 else ""
                    msg += (
                        "\n\n存在失败项，因此“上次成功获取时间”没有推进，下次可以安全重试。\n\n"
                        + detail
                    )
                    self.after(0, lambda: messagebox.showwarning(APP_TITLE, msg))
                else:
                    self.after(0, lambda: messagebox.showinfo(APP_TITLE, msg))
            except Exception as e:
                detail = traceback.format_exc()
                self.set_status("邮箱获取失败")
                self.after(
                    0,
                    lambda: messagebox.showerror(
                        APP_TITLE,
                        f"邮箱获取失败：{type(e).__name__}: {e}\n\n" + detail[-1800:],
                    ),
                )

        threading.Thread(target=worker, daemon=True).start()

    def _nutstore_credentials(self) -> tuple[str, str]:
        email = get_setting("nutstore_email", "").strip()
        enc = get_setting("nutstore_app_password", "").strip()
        if not email or not enc:
            raise ValueError("请先在“数据备份”中设置坚果云账号和应用密码")
        try:
            password = unprotect_secret(enc)
        except Exception as e:
            raise ValueError("已保存的坚果云应用密码无法解密，请重新输入") from e
        if not password:
            raise ValueError("请重新输入坚果云应用密码")
        return email, password

    def _perform_cloud_backup(self) -> str:
        email, password = self._nutstore_credentials()
        snapshot = temporary_snapshot_path()
        try:
            create_portable_snapshot(DB_PATH, snapshot)
            client = NutstoreWebDAV(email, password)
            history_name = client.upload_snapshot(snapshot, keep_history=30)
            set_setting("nutstore_last_backup", datetime.now().isoformat(timespec="seconds"))
            set_setting("nutstore_last_backup_error", "")
            return history_name
        finally:
            try:
                snapshot.unlink()
            except OSError:
                pass

    def schedule_auto_cloud_backup(self):
        if get_setting("nutstore_auto_backup", "1") != "1":
            return
        if not get_setting("nutstore_email", "").strip() or not get_setting("nutstore_app_password", "").strip():
            return
        if self._cloud_backup_running:
            self._cloud_backup_pending = True
            return
        self._cloud_backup_running = True

        def worker():
            try:
                name = self._perform_cloud_backup()
                self.set_status(f"坚果云自动备份完成：{name}")
            except Exception as e:
                error_text = f"{type(e).__name__}: {e}"
                try:
                    set_setting("nutstore_last_backup_error", error_text)
                except Exception:
                    pass
                self.set_status(f"坚果云自动备份失败：{e}")
            finally:
                self._cloud_backup_running = False
                if self._cloud_backup_pending:
                    self._cloud_backup_pending = False
                    self.after(100, self.schedule_auto_cloud_backup)

        threading.Thread(target=worker, daemon=True).start()

    def open_backup_dialog(self):
        win = tk.Toplevel(self)
        win.title("数据备份 / 恢复")
        win.transient(self)
        win.grab_set()
        win.geometry("650x520")
        outer = ttk.Frame(win, padding=14)
        outer.pack(fill="both", expand=True)

        ttk.Label(outer, text="坚果云 WebDAV", font=("TkDefaultFont", 10, "bold")).pack(anchor="w")
        ttk.Label(
            outer,
            text="备份内容：购买记录、人工确认、备注、显示设置等。PDF 发票本身不会上传。",
            wraplength=610,
        ).pack(anchor="w", pady=(4, 8))
        ttk.Label(
            outer,
            text="QQ/坚果云应用密码属于设备密钥，不写入云端备份；换电脑恢复后需要重新填写。",
            wraplength=610,
        ).pack(anchor="w", pady=(0, 10))

        email_var = tk.StringVar(value=get_setting("nutstore_email", ""))
        pass_var = tk.StringVar()
        auto_var = tk.BooleanVar(value=get_setting("nutstore_auto_backup", "1") == "1")
        ttk.Label(outer, text="坚果云账号邮箱").pack(anchor="w", pady=(6, 2))
        ttk.Entry(outer, textvariable=email_var).pack(fill="x")
        ttk.Label(outer, text="坚果云第三方应用密码（不是登录密码）").pack(anchor="w", pady=(8, 2))
        ttk.Entry(outer, textvariable=pass_var, show="•").pack(fill="x")
        has_saved = bool(get_setting("nutstore_app_password", ""))
        ttk.Label(
            outer,
            text=("已保存应用密码；留空继续使用。" if has_saved else "尚未保存应用密码。") + " Windows 下使用 DPAPI 加密保存。",
        ).pack(anchor="w", pady=(3, 6))
        ttk.Checkbutton(
            outer,
            text="数据变化后自动备份到坚果云（推荐）",
            variable=auto_var,
        ).pack(anchor="w", pady=(3, 8))

        last = get_setting("nutstore_last_backup", "")
        last_text = last.replace("T", " ")[:19] if last else "尚无成功备份"
        last_error = get_setting("nutstore_last_backup_error", "")
        status_var = tk.StringVar(
            value=f"最近成功备份：{last_text}" + (f"\n最近失败：{last_error}" if last_error else "")
        )
        ttk.Label(outer, textvariable=status_var, wraplength=610).pack(anchor="w", pady=(4, 10))
        ttk.Separator(outer).pack(fill="x", pady=8)

        def save_config(show_message=True):
            email = email_var.get().strip()
            if email and "@" not in email:
                messagebox.showwarning(APP_TITLE, "坚果云账号邮箱格式看起来不正确。", parent=win)
                return False
            try:
                with connect_db() as conn:
                    set_setting_conn(conn, "nutstore_email", email)
                    set_setting_conn(conn, "nutstore_auto_backup", "1" if auto_var.get() else "0")
                    if pass_var.get().strip():
                        set_setting_conn(conn, "nutstore_app_password", protect_secret(pass_var.get().strip()))
            except Exception as e:
                messagebox.showerror(APP_TITLE, f"保存坚果云设置失败：{type(e).__name__}: {e}", parent=win)
                return False
            if show_message:
                status_var.set("坚果云设置已保存。")
            return True

        def client_from_form() -> NutstoreWebDAV:
            email = email_var.get().strip()
            password = pass_var.get().strip()
            if not password:
                enc = get_setting("nutstore_app_password", "").strip()
                if enc:
                    password = unprotect_secret(enc)
            if not email or not password:
                raise ValueError("请填写坚果云账号和第三方应用密码")
            return NutstoreWebDAV(email, password)

        def test_connection():
            if not save_config(show_message=False):
                return
            status_var.set("正在测试坚果云连接...")

            def worker():
                try:
                    client_from_form().test_connection()
                    self.after(0, lambda: status_var.set("连接成功。InvoiceManager 备份目录已准备好。"))
                except Exception as e:
                    self.after(0, lambda: status_var.set(f"连接失败：{type(e).__name__}: {e}"))

            threading.Thread(target=worker, daemon=True).start()

        def backup_now():
            if not save_config(show_message=False):
                return
            status_var.set("正在备份到坚果云...")

            def worker():
                try:
                    name = self._perform_cloud_backup()
                    text = f"备份成功：{name}"
                    self.set_status(text)
                    self.after(0, lambda: status_var.set(text))
                except Exception as e:
                    error_text = f"{type(e).__name__}: {e}"
                    try:
                        set_setting("nutstore_last_backup_error", error_text)
                    except Exception:
                        pass
                    self.after(0, lambda: status_var.set(f"备份失败：{error_text}"))

            threading.Thread(target=worker, daemon=True).start()

        def restore_from_cloud():
            if not save_config(show_message=False):
                return
            status_var.set("正在读取坚果云历史备份...")

            def load_history_worker():
                try:
                    client = client_from_form()
                    history = client.list_history()
                    self.after(0, lambda: show_restore_choice(client, history))
                except Exception as e:
                    self.after(0, lambda: status_var.set(f"读取备份失败：{type(e).__name__}: {e}"))

            threading.Thread(target=load_history_worker, daemon=True).start()

        def show_restore_choice(client: NutstoreWebDAV, history: list[str]):
            choice_win = tk.Toplevel(win)
            choice_win.title("选择要恢复的备份")
            choice_win.transient(win)
            choice_win.grab_set()
            choice_win.geometry("560x220")
            box = ttk.Frame(choice_win, padding=14)
            box.pack(fill="both", expand=True)
            values = ["最新备份"] + history
            choice_var = tk.StringVar(value=values[0])
            ttk.Label(box, text="选择备份：").pack(anchor="w")
            ttk.Combobox(box, textvariable=choice_var, values=values, state="readonly").pack(fill="x", pady=(4, 10))
            ttk.Label(
                box,
                text="恢复会先在本机自动保存当前数据库，再替换为所选云端备份。恢复后程序会自动重启。",
                wraplength=520,
            ).pack(anchor="w", pady=(0, 10))

            def do_restore():
                selected = choice_var.get()
                remote_name = LATEST_NAME if selected == "最新备份" else selected
                if not messagebox.askyesno(
                    APP_TITLE,
                    "确定恢复这个备份？\n\n当前数据库会先自动做本地安全备份。",
                    parent=choice_win,
                ):
                    return
                choice_win.destroy()
                status_var.set("正在下载并校验备份...")

                def worker():
                    snapshot = temporary_snapshot_path(prefix="InvoiceManager-restore-")
                    try:
                        client.download_backup(snapshot, remote_name)
                        restore_snapshot(snapshot, DB_PATH, LOCAL_BACKUP_DIR)
                        hide_windows_file(DB_PATH)
                        hide_windows_file(LOCAL_BACKUP_DIR)
                        self.after(0, restart_after_restore)
                    except Exception as e:
                        self.after(0, lambda: status_var.set(f"恢复失败：{type(e).__name__}: {e}"))
                    finally:
                        try:
                            snapshot.unlink()
                        except OSError:
                            pass

                threading.Thread(target=worker, daemon=True).start()

            ttk.Button(box, text="恢复", command=do_restore).pack(side="right")
            ttk.Button(box, text="取消", command=choice_win.destroy).pack(side="right", padx=(0, 6))

        def restart_after_restore():
            messagebox.showinfo(
                APP_TITLE,
                "恢复完成。程序将重新启动。\n\n为保证跨电脑安全，QQ 邮箱授权码和坚果云应用密码不会从云备份恢复，需要重新填写。",
                parent=win,
            )
            try:
                subprocess.Popen([sys.executable], cwd=str(BASE_DIR))
            except Exception:
                pass
            self.destroy()

        def clear_password():
            if messagebox.askyesno(APP_TITLE, "确定清除已保存的坚果云应用密码？", parent=win):
                set_setting("nutstore_app_password", "")
                pass_var.set("")
                status_var.set("已清除坚果云应用密码。")

        row1 = ttk.Frame(outer)
        row1.pack(fill="x", pady=(4, 4))
        ttk.Button(row1, text="保存设置", command=save_config).pack(side="left")
        ttk.Button(row1, text="测试连接", command=test_connection).pack(side="left", padx=(6, 0))
        ttk.Button(row1, text="立即备份", command=backup_now).pack(side="left", padx=(6, 0))
        ttk.Button(row1, text="从坚果云恢复", command=restore_from_cloud).pack(side="left", padx=(6, 0))
        row2 = ttk.Frame(outer)
        row2.pack(fill="x", pady=(4, 0))
        ttk.Button(row2, text="清除应用密码", command=clear_password).pack(side="left")
        ttk.Button(row2, text="关闭", command=win.destroy).pack(side="right")

    def check_for_updates(self):
        self.set_status("正在检查更新...")

        def worker():
            try:
                info = latest_release(APP_VERSION)
                if not info.get("available"):
                    text = (
                        f"当前版本：v{APP_VERSION}\n"
                        + (f"最新正式版：v{info.get('latest')}\n\n当前已经是最新版。" if info.get("latest") else "暂无可用的正式 Release。")
                    )
                    self.set_status("检查更新完成")
                    self.after(0, lambda: messagebox.showinfo("检查更新", text))
                    return
                latest = info["latest"]
                release = info["release"]
                self.set_status(f"发现新版本 v{latest}")

                def ask():
                    if messagebox.askyesno(
                        "发现新版本",
                        f"当前版本：v{APP_VERSION}\n最新版本：v{latest}\n\n是否现在下载并更新？",
                    ):
                        self.download_and_install_update(release, latest)

                self.after(0, ask)
            except Exception as e:
                self.set_status("检查更新失败")
                self.after(0, lambda: messagebox.showerror("检查更新", f"检查更新失败：{type(e).__name__}: {e}"))

        threading.Thread(target=worker, daemon=True).start()

    def download_and_install_update(self, release: dict, latest: str):
        if not getattr(sys, "frozen", False):
            messagebox.showinfo(
                "更新",
                "当前正在以 Python 源码方式运行。自动替换只用于打包后的 Windows exe。",
            )
            return
        self.set_status(f"正在下载 v{latest}...")

        def worker():
            try:
                downloaded = download_verified_update(release, BASE_DIR)
                self.set_status("更新已下载并通过 SHA256 校验，正在重启...")

                def install():
                    try:
                        schedule_windows_replacement(Path(sys.executable), downloaded)
                    except Exception as e:
                        messagebox.showerror("更新失败", f"无法启动更新程序：{type(e).__name__}: {e}")
                        return
                    self.destroy()

                self.after(0, install)
            except Exception as e:
                self.set_status("下载更新失败")
                self.after(0, lambda: messagebox.showerror("更新失败", f"下载更新失败：{type(e).__name__}: {e}"))

        threading.Thread(target=worker, daemon=True).start()

    def do_export(self):
        try:
            count = export_excel(selected_fields())
            messagebox.showinfo(APP_TITLE, f"已导出 {count} 条发票记录：\n{EXPORT_PATH}")
            self.set_status(f"已导出 Excel：{EXPORT_FILENAME}")
        except PermissionError:
            messagebox.showerror(
                APP_TITLE,
                f"无法写入 {EXPORT_FILENAME}。\n请先关闭 Excel/WPS 中打开的同名文件。",
            )
        except Exception as e:
            messagebox.showerror(APP_TITLE, f"导出失败：{type(e).__name__}: {e}")


def main():
    InvoiceApp().mainloop()


if __name__ == "__main__":
    main()
