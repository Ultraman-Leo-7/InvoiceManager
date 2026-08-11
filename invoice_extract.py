# -*- coding: utf-8 -*-
"""
中国电子发票 PDF -> Excel 汇总（文件夹同步版）

核心行为：
1. 扫描脚本所在目录的所有 *.pdf。
2. Excel 与“当前文件夹”同步：
   - 新增 PDF -> 新增行
   - 删除 PDF -> 删除对应行
   - PDF 内容变化 -> 更新对应行
   - PDF 改名但内容不变 -> 更新文件名，不重新解析
3. 默认只显示：文件名、项目名称、价税合计。
4. settings.json 可配置显示字段；字段变化后 Excel 列会自动增删/重排。
5. 隐藏缓存保存在 Excel 内部，未变化的 PDF 不会重复解析。

依赖：
    pip install pymupdf openpyxl
"""

from __future__ import annotations

import hashlib
import json
import re
import sys
from copy import deepcopy
from pathlib import Path
from typing import Iterable

try:
    import fitz  # PyMuPDF
except ImportError:
    print("缺少依赖 pymupdf。请执行：pip install pymupdf openpyxl")
    raise SystemExit(1)

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("缺少依赖 openpyxl。请执行：pip install pymupdf openpyxl")
    raise SystemExit(1)


BASE_DIR = Path(__file__).resolve().parent
OUTPUT_XLSX = BASE_DIR / "发票汇总.xlsx"
SETTINGS_JSON = BASE_DIR / "settings.json"
CACHE_SHEET = "__invoice_cache__"

# 修改解析规则时增加此版本号，可让旧缓存自动失效并重新解析。
PARSER_VERSION = 4

DEFAULT_FIELDS = ["文件名", "项目名称", "价税合计"]

AVAILABLE_FIELDS = [
    "文件名",
    "发票号码",
    "开票日期",
    "购买方名称",
    "销售方名称",
    "项目名称",
    "价税合计",
    "金额（不含税）",
    "税额",
    "开票人",
    "提取状态",
]

INTERNAL_FIELDS = ["源文件SHA256", "解析版本"]

WIDTHS = {
    "文件名": 42,
    "发票号码": 24,
    "开票日期": 16,
    "购买方名称": 28,
    "销售方名称": 32,
    "项目名称": 46,
    "价税合计": 14,
    "金额（不含税）": 16,
    "税额": 12,
    "开票人": 12,
    "提取状态": 38,
}


# ---------- 设置 ----------

def write_default_settings() -> None:
    data = {
        "fields": DEFAULT_FIELDS,
        "available_fields": AVAILABLE_FIELDS,
    }
    SETTINGS_JSON.write_text(
        json.dumps(data, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def load_selected_fields() -> list[str]:
    if not SETTINGS_JSON.exists():
        write_default_settings()

    try:
        data = json.loads(SETTINGS_JSON.read_text(encoding="utf-8-sig"))
    except Exception as e:
        print(f"警告：settings.json 读取失败，将使用默认字段。原因：{e}")
        return DEFAULT_FIELDS.copy()

    raw = data.get("fields", DEFAULT_FIELDS)

    if isinstance(raw, str):
        if raw.strip().lower() in {"all", "*"}:
            return AVAILABLE_FIELDS.copy()
        raw = [raw]

    if not isinstance(raw, list):
        print("警告：settings.json 中 fields 必须是列表，将使用默认字段。")
        return DEFAULT_FIELDS.copy()

    result: list[str] = []
    invalid: list[str] = []
    for item in raw:
        name = str(item).strip()
        if name in AVAILABLE_FIELDS:
            if name not in result:
                result.append(name)
        elif name:
            invalid.append(name)

    if invalid:
        print("警告：忽略未知字段：" + "、".join(invalid))

    if not result:
        print("警告：没有有效显示字段，将使用默认字段。")
        return DEFAULT_FIELDS.copy()

    return result


# ---------- 通用工具 ----------

def norm(s: str | None) -> str:
    if not s:
        return ""
    s = s.replace("\u3000", " ").replace("\xa0", " ")
    s = re.sub(r"[ \t]+", " ", s)
    return s.strip()


def compact(s: str | None) -> str:
    return re.sub(r"\s+", "", s or "")


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def first_group(patterns: Iterable[str], text: str, flags: int = 0) -> str:
    for pattern in patterns:
        m = re.search(pattern, text, flags)
        if m:
            return norm(m.group(1))
    return ""


def money_to_float(s: str) -> float | None:
    if not s:
        return None
    s = s.replace(",", "").replace("￥", "").replace("¥", "").strip()
    try:
        return float(s)
    except ValueError:
        return None


# ---------- PDF 读取 ----------

def read_pdf(path: Path):
    doc = fitz.open(path)
    try:
        page_texts: list[str] = []
        pages_words: list[list[tuple]] = []
        for page in doc:
            page_texts.append(page.get_text("text", sort=True))
            pages_words.append(page.get_text("words", sort=True))
        return "\n".join(page_texts), pages_words
    finally:
        doc.close()


# ---------- 字段提取 ----------

def extract_party(text: str, party: str) -> str:
    t = compact(text)
    if party == "buyer":
        patterns = [
            r"购买方信息.*?名称[:：]?(.{2,80}?)(?=统一社会信用代码|纳税人识别号|销售方信息)",
            r"购买方.*?名称[:：]?(.{2,80}?)(?=统一社会信用代码|纳税人识别号|销售方)",
        ]
    else:
        patterns = [
            r"销售方信息.*?名称[:：]?(.{2,80}?)(?=统一社会信用代码|纳税人识别号|项目名称|规格型号)",
            r"销售方.*?名称[:：]?(.{2,80}?)(?=统一社会信用代码|纳税人识别号|项目名称|规格型号)",
        ]
    value = first_group(patterns, t, re.S)
    value = re.sub(r"^(?:名称[:：]?)", "", value)
    return value[:100]


def extract_items_from_text(text: str) -> list[str]:
    items: list[str] = []
    for raw in text.splitlines():
        line = norm(raw)
        if not line.startswith("*") or line.startswith("***"):
            continue
        name = line.split()[0] if " " in line else line
        name = norm(name)
        if len(name) >= 3 and name not in items:
            items.append(name)
    return items


def extract_items_from_coordinates(words_pages: list[list[tuple]]) -> list[str]:
    results: list[str] = []

    for words in words_pages:
        if not words:
            continue

        project_headers = [w for w in words if "项目名称" in str(w[4])]
        if not project_headers:
            continue

        h = project_headers[0]
        header_y = (h[1] + h[3]) / 2
        header_x = (h[0] + h[2]) / 2

        next_headers = [
            w for w in words
            if any(k in str(w[4]) for k in ("规格型号", "单位", "数量"))
            and ((w[0] + w[2]) / 2) > header_x
            and abs(((w[1] + w[3]) / 2) - header_y) < 25
        ]

        if next_headers:
            next_x = min((w[0] + w[2]) / 2 for w in next_headers)
            right_x = (header_x + next_x) / 2
        else:
            page_width = max(w[2] for w in words)
            right_x = page_width * 0.36

        total_candidates = [
            w for w in words
            if any(k in str(w[4]) for k in ("价税合计", "合计"))
            and ((w[1] + w[3]) / 2) > header_y + 15
        ]
        bottom_y = (
            min(((w[1] + w[3]) / 2) for w in total_candidates)
            if total_candidates
            else header_y + 220
        )

        candidates = [
            w for w in words
            if ((w[1] + w[3]) / 2) > header_y + 8
            and ((w[1] + w[3]) / 2) < bottom_y - 3
            and ((w[0] + w[2]) / 2) < right_x
            and str(w[4]).strip()
        ]

        rows: list[list[tuple]] = []
        for w in sorted(candidates, key=lambda x: (((x[1] + x[3]) / 2), x[0])):
            cy = (w[1] + w[3]) / 2
            if not rows:
                rows.append([w])
                continue
            prev_y = sum((x[1] + x[3]) / 2 for x in rows[-1]) / len(rows[-1])
            if abs(cy - prev_y) <= 4:
                rows[-1].append(w)
            else:
                rows.append([w])

        for row in rows:
            s = "".join(str(w[4]) for w in sorted(row, key=lambda x: x[0])).strip()
            if s.startswith("*") and len(s) >= 3 and s not in results:
                results.append(s)

    return results


def parse_invoice(path: Path) -> dict:
    text, words_pages = read_pdf(path)
    t = compact(text)

    if len(t) < 30:
        return {
            "文件名": path.name,
            "发票号码": "",
            "开票日期": "",
            "购买方名称": "",
            "销售方名称": "",
            "项目名称": "",
            "价税合计": None,
            "金额（不含税）": None,
            "税额": None,
            "开票人": "",
            "提取状态": "需复核：PDF 无可用文字层，可能是扫描件",
        }

    invoice_no = first_group([
        r"发票号码[:：]?\s*(\d{8,30})",
        r"发票号[:：]?\s*(\d{8,30})",
    ], t)

    issue_date = first_group([
        r"开票日期[:：]?\s*(\d{4}年\d{1,2}月\d{1,2}日)",
        r"开票日期[:：]?\s*(\d{4}[-/.]\d{1,2}[-/.]\d{1,2})",
    ], t)

    buyer = extract_party(text, "buyer")
    seller = extract_party(text, "seller")

    total_s = first_group([
        r"(?:价税合计)?[（(]?小写[）)]?[:：]?[￥¥]?\s*(-?[\d,]+\.\d{2})",
        r"价税合计.{0,120}?[￥¥]\s*(-?[\d,]+\.\d{2})",
    ], t, re.S)
    total = money_to_float(total_s)

    amount_wo_tax = None
    tax = None
    m = re.search(
        r"合计\s*[￥¥]?\s*(-?[\d,]+\.\d{2})\s*[￥¥]?\s*(-?[\d,]+\.\d{2})",
        compact(text),
    )
    if m:
        amount_wo_tax = money_to_float(m.group(1))
        tax = money_to_float(m.group(2))

    issuer = first_group([
        r"开票人[:：]?\s*([^\s]{1,30})",
    ], text)

    items = extract_items_from_coordinates(words_pages)
    if not items:
        items = extract_items_from_text(text)
    project = "；".join(items)

    warnings = []
    if not invoice_no:
        warnings.append("未识别发票号码")
    if not issue_date:
        warnings.append("未识别开票日期")
    if not project:
        warnings.append("未识别项目名称")
    if total is None:
        warnings.append("未识别价税合计")

    status = "正常" if not warnings else "需复核：" + "；".join(warnings)

    return {
        "文件名": path.name,
        "发票号码": invoice_no,
        "开票日期": issue_date,
        "购买方名称": buyer,
        "销售方名称": seller,
        "项目名称": project,
        "价税合计": total,
        "金额（不含税）": amount_wo_tax,
        "税额": tax,
        "开票人": issuer,
        "提取状态": status,
    }


# ---------- 缓存 ----------

CACHE_HEADERS = ["文件名", "源文件SHA256", "解析版本"] + [
    h for h in AVAILABLE_FIELDS if h != "文件名"
]


def load_cache():
    by_name: dict[str, dict] = {}
    by_hash: dict[str, dict] = {}

    if not OUTPUT_XLSX.exists():
        return by_name, by_hash

    try:
        wb = load_workbook(OUTPUT_XLSX, data_only=False, read_only=True)
    except Exception:
        return by_name, by_hash

    try:
        if CACHE_SHEET not in wb.sheetnames:
            # v3 及更早版本没有隐藏缓存；第一次运行 v4 时重新解析一次即可。
            return by_name, by_hash

        ws = wb[CACHE_SHEET]
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        header_map = {str(v): i + 1 for i, v in enumerate(headers) if v}

        for r in range(2, ws.max_row + 1):
            record: dict = {}
            for h in CACHE_HEADERS:
                c = header_map.get(h)
                record[h] = ws.cell(r, c).value if c else ""

            filename = str(record.get("文件名") or "").strip()
            digest = str(record.get("源文件SHA256") or "").strip()
            if not filename or not digest:
                continue

            by_name[filename] = record
            # 同内容文件可能改名；按 hash 也建立索引。
            by_hash.setdefault(digest, record)

        return by_name, by_hash
    finally:
        wb.close()


def cache_record_is_reusable(record: dict, digest: str) -> bool:
    try:
        version = int(record.get("解析版本") or 0)
    except Exception:
        version = 0
    return (
        str(record.get("源文件SHA256") or "") == digest
        and version == PARSER_VERSION
    )


# ---------- Excel 构建 ----------

def style_summary(ws, selected_fields: list[str]) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for i, h in enumerate(selected_fields, 1):
        ws.column_dimensions[get_column_letter(i)].width = WIDTHS.get(h, 18)

    for row in ws.iter_rows(min_row=2):
        for idx, cell in enumerate(row, 1):
            h = selected_fields[idx - 1]
            cell.alignment = Alignment(vertical="top", wrap_text=(h in {
                "文件名", "购买方名称", "销售方名称", "项目名称", "提取状态"
            }))
            if h in {"价税合计", "金额（不含税）", "税额"}:
                cell.number_format = "0.00"


def build_workbook(records: list[dict], selected_fields: list[str]) -> Workbook:
    wb = Workbook()

    ws = wb.active
    ws.title = "发票汇总"
    ws.append(selected_fields)

    for record in records:
        ws.append([record.get(h, "") for h in selected_fields])

    style_summary(ws, selected_fields)

    cache_ws = wb.create_sheet(CACHE_SHEET)
    cache_ws.append(CACHE_HEADERS)
    for record in records:
        cache_ws.append([record.get(h, "") for h in CACHE_HEADERS])
    cache_ws.sheet_state = "hidden"

    return wb


def atomic_save(wb: Workbook) -> None:
    tmp = OUTPUT_XLSX.with_name(OUTPUT_XLSX.stem + ".tmp.xlsx")
    if tmp.exists():
        try:
            tmp.unlink()
        except OSError:
            pass

    wb.save(tmp)
    try:
        tmp.replace(OUTPUT_XLSX)
    except PermissionError:
        try:
            tmp.unlink()
        except OSError:
            pass
        raise


# ---------- 主程序 ----------

def main() -> None:
    selected_fields = load_selected_fields()
    pdfs = sorted(BASE_DIR.glob("*.pdf"), key=lambda p: p.name.lower())

    old_by_name, old_by_hash = load_cache()

    records: list[dict] = []
    reused = 0
    parsed = 0
    failed = 0

    current_names = {p.name for p in pdfs}
    old_names = set(old_by_name.keys())
    removed = len(old_names - current_names)

    print(f"发票目录：{BASE_DIR}")
    print(f"当前 PDF：{len(pdfs)} 个")
    print("显示字段：" + "、".join(selected_fields))
    print("-" * 60)

    for i, pdf in enumerate(pdfs, 1):
        try:
            digest = sha256_file(pdf)

            old = old_by_name.get(pdf.name)
            if old and cache_record_is_reusable(old, digest):
                record = deepcopy(old)
                record["文件名"] = pdf.name
                reused += 1
                print(f"[{i}/{len(pdfs)}] 未变化：{pdf.name}")
                records.append(record)
                continue

            # 文件被改名但内容相同：复用旧解析结果。
            old_same_hash = old_by_hash.get(digest)
            if old_same_hash and cache_record_is_reusable(old_same_hash, digest):
                record = deepcopy(old_same_hash)
                record["文件名"] = pdf.name
                record["源文件SHA256"] = digest
                record["解析版本"] = PARSER_VERSION
                reused += 1
                print(f"[{i}/{len(pdfs)}] 已改名：{pdf.name}")
                records.append(record)
                continue

            record = parse_invoice(pdf)
            record["源文件SHA256"] = digest
            record["解析版本"] = PARSER_VERSION
            parsed += 1

            total = record.get("价税合计")
            total_text = f"{total:.2f}" if isinstance(total, (int, float)) else "?"
            print(
                f"[{i}/{len(pdfs)}] 已解析：{pdf.name} | "
                f"{record.get('项目名称') or '?'} | ¥{total_text}"
            )
            records.append(record)

        except Exception as e:
            failed += 1
            print(f"[{i}/{len(pdfs)}] 失败：{pdf.name} | {type(e).__name__}: {e}")
            # 解析异常不使用有效缓存版本，下次运行会继续尝试。
            records.append({
                "文件名": pdf.name,
                "发票号码": "",
                "开票日期": "",
                "购买方名称": "",
                "销售方名称": "",
                "项目名称": "",
                "价税合计": None,
                "金额（不含税）": None,
                "税额": None,
                "开票人": "",
                "提取状态": f"需复核：解析异常 {type(e).__name__}: {e}",
                "源文件SHA256": "",
                "解析版本": 0,
            })

    wb = build_workbook(records, selected_fields)

    try:
        atomic_save(wb)
    except PermissionError:
        print("\n保存失败：发票汇总.xlsx 可能正被 Excel/WPS 打开。")
        print("请关闭表格后重新运行脚本。")
        return
    finally:
        try:
            wb.close()
        except Exception:
            pass

    print("-" * 60)
    print(
        f"同步完成：当前 {len(records)} 条，"
        f"复用缓存 {reused}，重新解析 {parsed}，"
        f"删除旧记录 {removed}，异常 {failed}"
    )
    print(f"结果：{OUTPUT_XLSX}")


if __name__ == "__main__":
    main()
